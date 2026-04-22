"""Tests para backend/scripts/seed_supervisores.py

All DB and xlsx access is mocked — no real files or PostgreSQL required.
"""
from __future__ import annotations

import sys
import importlib
import pathlib
from unittest.mock import MagicMock, patch, call


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def _reload_seed():
    """Force-reload the module so state is fresh each test."""
    mod_name = "scripts.seed_supervisores"
    if mod_name in sys.modules:
        del sys.modules[mod_name]
    return importlib.import_module(mod_name)


def _make_cursor_mock():
    """Build a mock cursor that supports the context-manager protocol."""
    cur = MagicMock()
    ctx = MagicMock()
    ctx.__enter__ = MagicMock(return_value=cur)
    ctx.__exit__ = MagicMock(return_value=False)
    return ctx, cur


def _make_conn_mock():
    """Build a psycopg2 connection mock."""
    ctx, cur = _make_cursor_mock()
    conn = MagicMock()
    conn.cursor.return_value = ctx
    return conn, cur


# ---------------------------------------------------------------------------
# _load_rows_from_xlsx
# ---------------------------------------------------------------------------

def test_load_rows_reads_all_rows(monkeypatch, tmp_path):
    """_load_rows_from_xlsx must return all non-empty rows from the xlsx."""
    import openpyxl

    # Build a real xlsx in a temp dir
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["JUAN PEREZ", "SUPERVISOR_A"])
    ws.append(["MARIA GARCIA", "SUPERVISOR_B"])
    ws.append(["PEDRO LOPEZ", "SUPERVISOR_A"])
    path = tmp_path / "supervisores.xlsx"
    wb.save(str(path))

    seed = _reload_seed()
    rows = seed._load_rows_from_xlsx(path)

    assert len(rows) == 3
    assert rows[0] == ("JUAN PEREZ", "SUPERVISOR_A")
    assert rows[1] == ("MARIA GARCIA", "SUPERVISOR_B")
    assert rows[2] == ("PEDRO LOPEZ", "SUPERVISOR_A")


def test_load_rows_skips_empty_rows(monkeypatch, tmp_path):
    """Empty rows (None values) must be skipped."""
    import openpyxl

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["JUAN PEREZ", "SUPERVISOR_A"])
    ws.append([None, None])
    ws.append(["MARIA GARCIA", "SUPERVISOR_B"])
    path = tmp_path / "supervisores.xlsx"
    wb.save(str(path))

    seed = _reload_seed()
    rows = seed._load_rows_from_xlsx(path)

    assert len(rows) == 2


def test_load_rows_raises_if_file_missing():
    """FileNotFoundError must be raised when the xlsx path doesn't exist."""
    import pytest

    seed = _reload_seed()
    with pytest.raises(FileNotFoundError):
        seed._load_rows_from_xlsx(pathlib.Path("/nonexistent/supervisores.xlsx"))


def test_load_rows_raises_if_empty_file(tmp_path):
    """ValueError must be raised when xlsx has no data rows."""
    import openpyxl
    import pytest

    wb = openpyxl.Workbook()
    path = tmp_path / "empty.xlsx"
    wb.save(str(path))

    seed = _reload_seed()
    with pytest.raises(ValueError, match="no contiene filas"):
        seed._load_rows_from_xlsx(path)


def test_load_rows_strips_whitespace(tmp_path):
    """Leading/trailing whitespace in cell values must be stripped."""
    import openpyxl

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["  JUAN PEREZ  ", "  SUPERVISOR_A  "])
    path = tmp_path / "supervisores.xlsx"
    wb.save(str(path))

    seed = _reload_seed()
    rows = seed._load_rows_from_xlsx(path)

    assert rows[0] == ("JUAN PEREZ", "SUPERVISOR_A")


# ---------------------------------------------------------------------------
# seed_supervisores — DB interactions
# ---------------------------------------------------------------------------

def _fake_rows():
    return [
        ("JUAN PEREZ", "SUPERVISOR_A"),
        ("MARIA GARCIA", "SUPERVISOR_B"),
    ]


def test_seed_upserts_each_row(monkeypatch):
    """Each row must trigger an INSERT ... ON CONFLICT upsert."""
    monkeypatch.setattr("dotenv.load_dotenv", lambda *a, **kw: None)

    seed = _reload_seed()
    conn, cur = _make_conn_mock()

    executed_sqls: list[tuple[str, tuple]] = []

    def capture_execute(sql, params=None):
        executed_sqls.append((sql.strip(), params))

    cur.execute.side_effect = capture_execute

    with (
        patch("scripts.seed_supervisores.psycopg2.connect", return_value=conn),
        patch.object(seed, "_load_rows_from_xlsx", return_value=_fake_rows()),
    ):
        seed.seed_supervisores()

    upsert_sqls = [
        (sql, params) for sql, params in executed_sqls
        if "INSERT INTO operations.preventistas_supervisores" in sql
    ]
    assert len(upsert_sqls) == 2, (
        f"Expected 2 upsert calls, got {len(upsert_sqls)}. "
        f"All SQL: {[s for s, _ in executed_sqls]}"
    )


def test_seed_uses_on_conflict_for_idempotency(monkeypatch):
    """All upserts must include ON CONFLICT DO UPDATE for idempotency."""
    monkeypatch.setattr("dotenv.load_dotenv", lambda *a, **kw: None)

    seed = _reload_seed()
    conn, cur = _make_conn_mock()

    executed_sqls: list[str] = []
    cur.execute.side_effect = lambda sql, params=None: executed_sqls.append(sql.strip())

    with (
        patch("scripts.seed_supervisores.psycopg2.connect", return_value=conn),
        patch.object(seed, "_load_rows_from_xlsx", return_value=_fake_rows()),
    ):
        seed.seed_supervisores()

    insert_sqls = [s for s in executed_sqls if "INSERT" in s.upper()]
    assert insert_sqls, "No INSERT SQL found"
    for sql in insert_sqls:
        assert "ON CONFLICT" in sql.upper(), (
            f"INSERT must use ON CONFLICT for idempotency. Got: {sql}"
        )
        assert "DO UPDATE" in sql.upper(), (
            f"INSERT must use DO UPDATE for idempotency. Got: {sql}"
        )


def test_seed_commits_transaction(monkeypatch):
    """Connection must be committed after all upserts."""
    monkeypatch.setattr("dotenv.load_dotenv", lambda *a, **kw: None)

    seed = _reload_seed()
    conn, cur = _make_conn_mock()

    with (
        patch("scripts.seed_supervisores.psycopg2.connect", return_value=conn),
        patch.object(seed, "_load_rows_from_xlsx", return_value=_fake_rows()),
    ):
        seed.seed_supervisores()

    conn.commit.assert_called_once()


def test_seed_exits_1_if_file_missing(monkeypatch):
    """seed_supervisores must exit with code 1 when xlsx is missing."""
    import pytest

    monkeypatch.setattr("dotenv.load_dotenv", lambda *a, **kw: None)

    seed = _reload_seed()

    with pytest.raises(SystemExit) as exc_info:
        seed.seed_supervisores(xlsx_path=pathlib.Path("/nonexistent/supervisores.xlsx"))

    assert exc_info.value.code == 1


def test_seed_correct_params_passed_to_upsert(monkeypatch):
    """Upsert must pass (preventista, supervisor) as params in correct order."""
    monkeypatch.setattr("dotenv.load_dotenv", lambda *a, **kw: None)

    seed = _reload_seed()
    conn, cur = _make_conn_mock()

    captured_params: list[tuple] = []

    def capture_execute(sql, params=None):
        if "INSERT" in sql.upper():
            captured_params.append(params)

    cur.execute.side_effect = capture_execute

    single_row = [("EZEQUIEL CACHAGUA", "GFLORES")]

    with (
        patch("scripts.seed_supervisores.psycopg2.connect", return_value=conn),
        patch.object(seed, "_load_rows_from_xlsx", return_value=single_row),
    ):
        seed.seed_supervisores()

    assert len(captured_params) == 1
    params = captured_params[0]
    assert params[0] == "EZEQUIEL CACHAGUA", f"First param must be preventista, got: {params[0]}"
    assert params[1] == "GFLORES", f"Second param must be supervisor, got: {params[1]}"


def test_seed_idempotent_on_double_run(monkeypatch):
    """Running the seed twice must not raise errors."""
    monkeypatch.setattr("dotenv.load_dotenv", lambda *a, **kw: None)

    seed = _reload_seed()
    conn, cur = _make_conn_mock()

    with (
        patch("scripts.seed_supervisores.psycopg2.connect", return_value=conn),
        patch.object(seed, "_load_rows_from_xlsx", return_value=_fake_rows()),
    ):
        seed.seed_supervisores()
        seed.seed_supervisores()

    # Both runs must commit
    assert conn.commit.call_count == 2
