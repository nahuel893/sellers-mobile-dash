"""Seed script: ingest visitados.xlsx into operations.visitas_preventista.

Usage (from backend/):
    .venv/bin/python -m scripts.seed_visitas
    .venv/bin/python -m scripts.seed_visitas /path/to/visitados.xlsx

Configuration via environment variables:
    VISITAS_XLSX_PATH  — path to the Excel file
                         (default: project-root/../sales-dashboard/data/visitados.xlsx)

PostgreSQL connection uses the same env vars as the main backend:
    DB_HOST, DB_PORT, DB_NAME, DB_USER, DB_PASSWORD

The script is importable for testing (see tests/test_seed_visitas.py).
The main entry-point is ``seed_visitas(xlsx_path)`` which can be called directly.

Excel column layout (Hoja1):
    col 0  — ignored (empty)
    col 1  — Id cliente        → id_cliente (INT)
    col 2  — Id cliente erp    → ignored
    col 3  — Descripción cliente → descripcion_cliente (TEXT)
    col 4  — Id ruta           → id_ruta (INT)
    col 5  — Ruta              → ruta (TEXT)
    col 6  — Domicilio         → domicilio (TEXT)
    col 7  — Sector            → sector (TEXT)
    col 8  — Fecha             → fecha (DATE, format DD-MM-YYYY or datetime)
    col 9  — Visitado          → visitado (BOOLEAN, 'SI'/'NO')
    col 10 — Hora visita       → hora_visita (TIME or str HH:MM:SS)
    col 11 — Hora venta        → hora_venta (TIME or str)
    col 12 — Hora motivo       → hora_motivo (TIME or str)
    col 13 — Motivo            → motivo (TEXT)
    col 14 — Latitud           → latitud (FLOAT)
    col 15 — Longitud          → longitud (FLOAT)
    col 16 — f/h foto          → foto_url (TEXT)
"""
from __future__ import annotations

import datetime
import os
import pathlib
import sys
from typing import Any

import openpyxl
import psycopg2
from dotenv import load_dotenv

# Load .env from project root (3 levels above backend/scripts/seed_visitas.py)
_env_path = pathlib.Path(__file__).resolve().parent.parent.parent / ".env"
load_dotenv(_env_path)

# ---------------------------------------------------------------------------
# Constants
# ---------------------------------------------------------------------------

_DEFAULT_XLSX_PATH = (
    pathlib.Path(__file__).resolve().parent.parent.parent.parent
    / "sales-dashboard"
    / "data"
    / "visitados.xlsx"
)

# Required columns (by index in the Excel sheet, 0-based)
_COL_ID_CLIENTE = 1
_COL_DESCRIPCION = 3
_COL_ID_RUTA = 4
_COL_RUTA = 5
_COL_DOMICILIO = 6
_COL_SECTOR = 7
_COL_FECHA = 8
_COL_VISITADO = 9
_COL_HORA_VISITA = 10
_COL_HORA_VENTA = 11
_COL_HORA_MOTIVO = 12
_COL_MOTIVO = 13
_COL_LATITUD = 14
_COL_LONGITUD = 15
_COL_FOTO_URL = 16

_MIN_COLUMNS = 16  # minimum number of columns expected in each data row


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def _parse_date(raw: Any) -> datetime.date | None:
    """Parse a date value from the Excel cell (string 'DD-MM-YYYY' or datetime)."""
    if raw is None:
        return None
    if isinstance(raw, (datetime.date, datetime.datetime)):
        return raw.date() if isinstance(raw, datetime.datetime) else raw
    if isinstance(raw, str):
        raw = raw.strip()
        if not raw:
            return None
        # Try DD-MM-YYYY (legacy format used in visitados.xlsx)
        try:
            return datetime.datetime.strptime(raw, "%d-%m-%Y").date()
        except ValueError:
            pass
        # Fallback: YYYY-MM-DD
        try:
            return datetime.datetime.strptime(raw, "%Y-%m-%d").date()
        except ValueError:
            pass
    return None


def _parse_time(raw: Any) -> datetime.time | None:
    """Parse a time value from the Excel cell (string 'HH:MM:SS' or time)."""
    if raw is None:
        return None
    if isinstance(raw, datetime.time):
        return raw
    if isinstance(raw, datetime.datetime):
        return raw.time()
    if isinstance(raw, str):
        raw = raw.strip()
        if not raw:
            return None
        try:
            return datetime.time.fromisoformat(raw)
        except ValueError:
            pass
    return None


def _parse_bool(raw: Any) -> bool:
    """Parse visitado flag: 'SI' → True, anything else → False."""
    if raw is None:
        return False
    if isinstance(raw, bool):
        return raw
    return str(raw).strip().upper() == "SI"


def _parse_float(raw: Any) -> float | None:
    """Parse a float, treating 0 as None (coordinates: 0 is invalid)."""
    if raw is None:
        return None
    try:
        v = float(raw)
        return v if v != 0.0 else None
    except (TypeError, ValueError):
        return None


def _parse_int(raw: Any) -> int | None:
    """Parse an integer or return None."""
    if raw is None:
        return None
    try:
        return int(raw)
    except (TypeError, ValueError):
        return None


def _row_to_record(row: tuple) -> dict | None:
    """Convert a raw Excel row tuple to an insert record dict.

    Returns None if the row is missing required fields (id_cliente or fecha).
    """
    if len(row) <= _MIN_COLUMNS:
        return None

    id_cliente = _parse_int(row[_COL_ID_CLIENTE])
    fecha = _parse_date(row[_COL_FECHA])

    # Skip rows without the mandatory fields
    if id_cliente is None or fecha is None:
        return None

    return {
        "id_cliente": id_cliente,
        "id_ruta": _parse_int(row[_COL_ID_RUTA]),
        "ruta": str(row[_COL_RUTA]).strip() if row[_COL_RUTA] is not None else None,
        "fecha": fecha,
        "hora_visita": _parse_time(row[_COL_HORA_VISITA]),
        "hora_venta": _parse_time(row[_COL_HORA_VENTA]),
        "hora_motivo": _parse_time(row[_COL_HORA_MOTIVO]),
        "motivo": str(row[_COL_MOTIVO]).strip() if row[_COL_MOTIVO] is not None else None,
        "visitado": _parse_bool(row[_COL_VISITADO]),
        "latitud": _parse_float(row[_COL_LATITUD]),
        "longitud": _parse_float(row[_COL_LONGITUD]),
        "domicilio": str(row[_COL_DOMICILIO]).strip() if row[_COL_DOMICILIO] is not None else None,
        "sector": str(row[_COL_SECTOR]).strip() if row[_COL_SECTOR] is not None else None,
        "descripcion_cliente": str(row[_COL_DESCRIPCION]).strip() if row[_COL_DESCRIPCION] is not None else None,
        "foto_url": str(row[_COL_FOTO_URL]).strip() if row[_COL_FOTO_URL] is not None else None,
    }


# ---------------------------------------------------------------------------
# Core function
# ---------------------------------------------------------------------------

def seed_visitas(xlsx_path: str | pathlib.Path | None = None) -> dict[str, int]:
    """Ingest visitados.xlsx into operations.visitas_preventista (idempotent).

    Args:
        xlsx_path: Path to the Excel file. Falls back to VISITAS_XLSX_PATH env
                   var, then the default path relative to the legacy project.

    Returns:
        dict with keys 'inserted' and 'skipped'.

    Raises:
        FileNotFoundError: if the xlsx file does not exist.
        ValueError: if required columns are missing from the file.
        SystemExit: on DB connection failure (mirrors seed_admin pattern).
    """
    # Resolve path
    if xlsx_path is None:
        xlsx_path = os.getenv("VISITAS_XLSX_PATH", str(_DEFAULT_XLSX_PATH))
    xlsx_path = pathlib.Path(xlsx_path)

    if not xlsx_path.exists():
        print(f"ERROR: Excel file not found: {xlsx_path}", file=sys.stderr)
        sys.exit(1)

    # Read Excel
    print(f"Reading {xlsx_path} …")
    wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
    ws = wb.active

    # Validate headers (first row)
    header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), None)
    if header_row is None:
        print("ERROR: Excel file is empty.", file=sys.stderr)
        sys.exit(1)

    # Build records from data rows (skip header)
    records = []
    skipped_parse = 0
    for row in ws.iter_rows(min_row=2, values_only=True):
        rec = _row_to_record(row)
        if rec is None:
            skipped_parse += 1
        else:
            records.append(rec)
    wb.close()

    print(f"Parsed {len(records)} valid rows ({skipped_parse} skipped — missing id_cliente or fecha).")

    if not records:
        print("No records to insert. Exiting.")
        return {"inserted": 0, "skipped": 0}

    # Connect to PostgreSQL
    conn = psycopg2.connect(
        host=os.getenv("DB_HOST", "localhost"),
        port=os.getenv("DB_PORT", "5432"),
        dbname=os.getenv("DB_NAME", "gold_db"),
        user=os.getenv("DB_USER"),
        password=os.getenv("DB_PASSWORD"),
    )

    inserted = 0
    skipped_conflict = 0

    insert_sql = """
        INSERT INTO operations.visitas_preventista (
            id_cliente, id_ruta, ruta, fecha,
            hora_visita, hora_venta, hora_motivo, motivo,
            visitado, latitud, longitud,
            domicilio, sector, descripcion_cliente, foto_url
        ) VALUES (
            %(id_cliente)s, %(id_ruta)s, %(ruta)s, %(fecha)s,
            %(hora_visita)s, %(hora_venta)s, %(hora_motivo)s, %(motivo)s,
            %(visitado)s, %(latitud)s, %(longitud)s,
            %(domicilio)s, %(sector)s, %(descripcion_cliente)s, %(foto_url)s
        )
        ON CONFLICT (id_cliente, fecha, hora_visita) DO NOTHING
    """

    with conn.cursor() as cur:
        for rec in records:
            cur.execute(insert_sql, rec)
            if cur.rowcount == 1:
                inserted += 1
            else:
                skipped_conflict += 1

    conn.commit()
    conn.close()

    total_skipped = skipped_parse + skipped_conflict
    print(
        f"Done. Inserted: {inserted} | Skipped (conflict): {skipped_conflict} "
        f"| Skipped (parse): {skipped_parse}"
    )
    return {"inserted": inserted, "skipped": total_skipped}


# ---------------------------------------------------------------------------
# CLI entry-point
# ---------------------------------------------------------------------------

if __name__ == "__main__":
    path_arg = sys.argv[1] if len(sys.argv) > 1 else None
    try:
        result = seed_visitas(path_arg)
        sys.exit(0)
    except Exception as exc:  # noqa: BLE001
        print(f"ERROR: {exc}", file=sys.stderr)
        sys.exit(1)
